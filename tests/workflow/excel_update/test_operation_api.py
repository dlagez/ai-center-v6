from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _build_duplicate_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["项目编号", "项目名称", "回款"])
    worksheet.append(["P001", "项目A", 10])
    worksheet.append(["P002", "项目B", 20])
    worksheet.append(["P002", "项目B-重复", 30])
    workbook.save(path)


def _build_simple_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["项目编号", "项目名称", "回款"])
    worksheet.append(["P001", "项目A", 10])
    worksheet.append(["P002", "项目B", 20])
    workbook.save(path)


def _relative_path(url: str) -> str:
    return url.replace("http://testserver", "")


def _wait_for_terminal_status(client, task_id: int, operation_id: int) -> dict:
    for _ in range(10):
        response = client.get(f"/api/excel-tasks/{task_id}/operations/{operation_id}/status")
        payload = response.json()
        status = payload["data"]["status"]
        if status not in {"pending", "running"}:
            return payload
    return payload


def test_create_operation_generates_new_version_after_successful_updates(client, excel_env: Path) -> None:
    excel_path = excel_env / "duplicate.xlsx"
    _build_duplicate_workbook(excel_path)

    with excel_path.open("rb") as file_obj:
        create_task_response = client.post(
            "/api/excel-tasks/upload",
            files={
                "file": (
                    "duplicate.xlsx",
                    file_obj,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    task_id = create_task_response.json()["data"]["task_id"]
    create_operation_response = client.post(
        f"/api/excel-tasks/{task_id}/operations",
        json={
            "sheet_name": "Sheet1",
            "match_column": "A",
            "target_column": "C",
            "submitted_by": "lisi",
            "updates": [
                {"match_value": "P001", "target_value": 99},
                {"match_value": "P002", "target_value": 77},
                {"match_value": "P404", "target_value": 12},
            ],
        },
    )

    assert create_operation_response.status_code == 200
    create_payload = create_operation_response.json()
    assert create_payload["code"] == 0
    assert create_payload["data"]["status"] == "pending"

    operation_id = create_payload["data"]["operation_id"]
    status_payload = _wait_for_terminal_status(client, task_id, operation_id)
    assert status_payload["data"]["status"] == "partial_success"
    assert status_payload["data"]["summary"] == {
        "total_count": 3,
        "success_count": 1,
        "not_found_count": 1,
        "duplicate_count": 1,
        "failed_count": 0,
    }

    detail_response = client.get(f"/api/excel-tasks/{task_id}/operations/{operation_id}")
    detail_payload = detail_response.json()
    assert detail_payload["data"]["submitted_by"] == "lisi"
    assert detail_payload["data"]["result_version"]["version_no"] == 2
    assert [item["status"] for item in detail_payload["data"]["items"]] == [
        "success",
        "duplicate_match",
        "not_found",
    ]

    version_response = client.get(f"/api/excel-tasks/{task_id}/versions")
    version_payload = version_response.json()
    assert version_payload["data"]["total"] == 2
    assert version_payload["data"]["items"][0]["version_no"] == 2
    assert version_payload["data"]["items"][0]["is_current"] is True

    latest_download_response = client.get(f"/api/excel-tasks/{task_id}/download/latest")
    latest_download_url = latest_download_response.json()["data"]["download_url"]
    file_response = client.get(_relative_path(latest_download_url))
    workbook = load_workbook(filename=BytesIO(file_response.content))
    worksheet = workbook["Sheet1"]
    assert worksheet["C2"].value == 99
    assert worksheet["C3"].value == 20
    assert worksheet["C4"].value == 30


def test_create_operation_without_success_does_not_generate_new_version(client, excel_env: Path) -> None:
    excel_path = excel_env / "simple.xlsx"
    _build_simple_workbook(excel_path)

    with excel_path.open("rb") as file_obj:
        create_task_response = client.post(
            "/api/excel-tasks/upload",
            files={
                "file": (
                    "simple.xlsx",
                    file_obj,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    task_id = create_task_response.json()["data"]["task_id"]
    create_operation_response = client.post(
        f"/api/excel-tasks/{task_id}/operations",
        json={
            "sheet_name": "Sheet1",
            "match_column": "A",
            "target_column": "C",
            "updates": [
                {"match_value": "P404", "target_value": 99},
                {"match_value": "P405", "target_value": 88},
            ],
        },
    )

    operation_id = create_operation_response.json()["data"]["operation_id"]
    status_payload = _wait_for_terminal_status(client, task_id, operation_id)
    assert status_payload["data"]["status"] == "partial_success"
    assert status_payload["data"]["summary"]["success_count"] == 0

    detail_response = client.get(f"/api/excel-tasks/{task_id}/operations/{operation_id}")
    detail_payload = detail_response.json()
    assert detail_payload["data"]["result_version"] is None
    assert detail_payload["data"]["summary"]["not_found_count"] == 2

    version_response = client.get(f"/api/excel-tasks/{task_id}/versions")
    assert version_response.json()["data"]["total"] == 1
