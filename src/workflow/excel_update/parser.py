from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.workflow.excel_update.schemas import ExcelUpdateRequest


def _normalize_header_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "".join(text.split())


def parse_excel_template(request: ExcelUpdateRequest) -> dict[str, Any]:
    excel_path = Path(request.excel_path)
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel file not found: {request.excel_path}")

    workbook = load_workbook(excel_path)
    if request.sheet_name:
        if request.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {request.sheet_name}")
        worksheet = workbook[request.sheet_name]
    else:
        worksheet = workbook.active

    header_row_index = None
    match_column_index = None
    target_column_index = None
    match_column_row_index = None
    target_column_row_index = None
    max_scan_rows = min(worksheet.max_row, 10)
    normalized_match_column = _normalize_header_text(request.match_column)
    normalized_target_column = _normalize_header_text(request.target_column)

    for row_index in range(1, max_scan_rows + 1):
        row_values = {
            column_index: _normalize_header_text(cell_value)
            for column_index, cell_value in enumerate(
                next(
                    worksheet.iter_rows(
                        min_row=row_index,
                        max_row=row_index,
                        values_only=True,
                    )
                ),
                start=1,
            )
            if cell_value not in (None, "")
        }

        if match_column_index is None:
            current_match_column = next(
                (
                    column_index
                    for column_index, value in row_values.items()
                    if value == normalized_match_column
                ),
                None,
            )
            if current_match_column is not None:
                match_column_index = current_match_column
                match_column_row_index = row_index

        if target_column_index is None:
            current_target_column = next(
                (
                    column_index
                    for column_index, value in row_values.items()
                    if value == normalized_target_column
                ),
                None,
            )
            if current_target_column is not None:
                target_column_index = current_target_column
                target_column_row_index = row_index

        if match_column_index is not None and target_column_index is not None:
            break

    if (
        match_column_index is None
        or target_column_index is None
        or match_column_row_index is None
        or target_column_row_index is None
    ):
        raise ValueError(
            f"Unable to locate header row with match_column='{request.match_column}' "
            f"and target_column='{request.target_column}'"
        )

    header_row_index = max(match_column_row_index, target_column_row_index)

    row_index_by_match_value: dict[str, int] = {}
    duplicate_match_values: list[str] = []
    data_start_row = header_row_index + 1
    while data_start_row <= worksheet.max_row:
        cell_value = worksheet.cell(row=data_start_row, column=match_column_index).value
        if cell_value not in (None, ""):
            break
        data_start_row += 1

    for row_index in range(data_start_row, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row_index, column=match_column_index).value
        if cell_value in (None, ""):
            continue
        match_value = str(cell_value).strip()
        if not match_value:
            continue
        if match_value in row_index_by_match_value:
            duplicate_match_values.append(match_value)
            continue
        row_index_by_match_value[match_value] = row_index

    return {
        "excel_path": request.excel_path,
        "sheet_name": worksheet.title,
        "match_column": request.match_column,
        "match_field": request.match_field,
        "target_column": request.target_column,
        "header_row_index": header_row_index,
        "data_start_row": data_start_row,
        "match_column_index": match_column_index,
        "target_column_index": target_column_index,
        "row_index_by_match_value": row_index_by_match_value,
        "duplicate_match_values": duplicate_match_values,
        "workbook": workbook,
        "worksheet": worksheet,
    }
