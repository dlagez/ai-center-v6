from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.workflow.excel_update.schemas import ExcelUpdateRequest


def _normalize_header_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "".join(text.split())


def fetch_excel_records(request: ExcelUpdateRequest) -> list[dict[str, Any]]:
    source_excel_path = request.source_excel_path
    if not source_excel_path:
        raise ValueError("source_excel_path is required when source_type is excel_file")

    excel_path = Path(source_excel_path)
    if not excel_path.is_file():
        raise FileNotFoundError(f"Source Excel file not found: {source_excel_path}")

    workbook = load_workbook(excel_path, data_only=True)
    if request.source_sheet_name:
        if request.source_sheet_name not in workbook.sheetnames:
            raise ValueError(f"Source sheet not found: {request.source_sheet_name}")
        worksheet = workbook[request.source_sheet_name]
    else:
        worksheet = workbook.active

    source_match_column = request.source_match_column or request.match_column
    source_value_column = request.source_value_column or request.target_column
    normalized_match = _normalize_header_text(source_match_column)
    normalized_value = _normalize_header_text(source_value_column)

    match_column_index = None
    value_column_index = None
    max_scan_rows = min(worksheet.max_row, 10)

    for row_index in range(1, max_scan_rows + 1):
        row_values = {
            column_index: _normalize_header_text(cell_value)
            for column_index, cell_value in enumerate(
                next(
                    worksheet.iter_rows(
                        min_row=row_index,
                        max_row=row_index,
                        values_only=True,
                    )
                ),
                start=1,
            )
            if cell_value not in (None, "")
        }

        if match_column_index is None:
            match_column_index = next(
                (column_index for column_index, value in row_values.items() if value == normalized_match),
                None,
            )
        if value_column_index is None:
            value_column_index = next(
                (column_index for column_index, value in row_values.items() if value == normalized_value),
                None,
            )
        if match_column_index is not None and value_column_index is not None:
            header_row_index = row_index
            break
    else:
        raise ValueError(
            f"Unable to locate source columns match_column='{source_match_column}' "
            f"and value_column='{source_value_column}'"
        )

    records: list[dict[str, Any]] = []
    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        match_value = worksheet.cell(row=row_index, column=match_column_index).value
        value = worksheet.cell(row=row_index, column=value_column_index).value
        if match_value in (None, "") or value in (None, ""):
            continue
        records.append(
            {
                request.match_field: str(match_value).strip(),
                "value": value,
            }
        )

    workbook.close()
    return records
