from datetime import datetime
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from src.api.app import app
from src.workflow.excel_update.analyzer import analyze_excel_update


client = TestClient(app)


def _build_target_excel(path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "封面"
    summary_sheet.append(["说明", "请勿修改"])

    worksheet = workbook.create_sheet("软件项目清欠表_收付款")
    worksheet.append(["说明", "", "", "", ""])
    worksheet.append(["序号", "项目编号", "项目名称", "3月实际产值", "3月回款"])
    worksheet.append(["1", "HKZC-N-YW-2021-001", "项目A", None, None])
    worksheet.append(["2", "HKZC-N-YW-2021-002", "项目B", None, None])
    workbook.save(path)


def _build_source_excel(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "数据源"
    worksheet.append(["序号", "月份", "项目编号", "交易时间", "户名", "回款金额", "项目名称"])
    worksheet.append([12, "2月", "HKZC-N-YW-2021-001", "2026/1/30", "A公司", 559336.13, "项目A"])
    worksheet.append([13, "2月", "HKZC-N-YW-2021-002", "2026/1/31", "B公司", 223344.55, "项目B"])
    workbook.save(path)


def test_excel_update_analysis_infers_sheet_and_target_column(tmp_path) -> None:
    excel_path = tmp_path / "analysis.xlsx"
    _build_target_excel(excel_path)

    result = analyze_excel_update(str(excel_path), "把这张清欠表 3 月实际产值更新一下")

    assert result.sheet_name == "软件项目清欠表_收付款"
    assert result.match_column == "项目编号"
    assert result.target_column == "3月实际产值"
    assert [item.model_dump(mode="json") for item in result.query_conditions] == [
        {"key": "month", "value": f"{datetime.now().year:04d}-03"},
        {"key": "metric", "value": "actual_output"},
    ]


def test_excel_update_analysis_can_identify_source_excel_columns(tmp_path) -> None:
    target_excel_path = tmp_path / "target.xlsx"
    source_excel_path = tmp_path / "source.xlsx"
    _build_target_excel(target_excel_path)
    _build_source_excel(source_excel_path)

    result = analyze_excel_update(
        str(target_excel_path),
        "使用源excel的项目编号、回款金额，来更新清欠表的三月回款",
        source_excel_path=str(source_excel_path),
    )

    assert result.sheet_name == "软件项目清欠表_收付款"
    assert result.target_column == "3月回款"
    assert result.source_match_column == "项目编号"
    assert result.source_value_column == "回款金额"


def test_task_supports_multiple_operations_after_single_upload(tmp_path) -> None:
    excel_path = tmp_path / "analysis.xlsx"
    _build_target_excel(excel_path)

    with excel_path.open("rb") as file_obj:
        create_response = client.post(
            "/workflow/excel-update/tasks",
            files={"file": ("analysis.xlsx", file_obj, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert create_response.status_code == 200
    task_payload = create_response.json()
    task_id = task_payload["task_id"]
    assert task_payload["operation_count"] == 0

    operation_1 = client.post(
        f"/workflow/excel-update/tasks/{task_id}/operations",
        data={"user_prompt": "把这张清欠表 3 月实际产值更新一下"},
    )
    assert operation_1.status_code == 200
    payload_1 = operation_1.json()
    assert payload_1["sequence"] == 1
    assert payload_1["request"]["target_column"] == "3月实际产值"
    assert payload_1["result"]["summary"]["updated_cells"] == 2

    operation_2 = client.post(
        f"/workflow/excel-update/tasks/{task_id}/operations",
        data={
            "user_prompt": "把这张清欠表 3 月回款更新一下",
            "query_conditions": '[{"key":"month","value":"2026-03"},{"key":"metric","value":"receipts"}]',
        },
    )
    assert operation_2.status_code == 200
    payload_2 = operation_2.json()
    assert payload_2["sequence"] == 2
    assert payload_2["request"]["target_column"] == "3月回款"

    detail_response = client.get(f"/workflow/excel-update/tasks/{task_id}")
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["operation_count"] == 2
    assert len(detail["operations"]) == 2
    assert detail["latest_target_column"] == "3月回款"

    workbook = load_workbook(detail["current_excel_path"])
    worksheet = workbook["软件项目清欠表_收付款"]
    assert worksheet["D3"].value == 20
    assert worksheet["D4"].value == 30
    assert worksheet["E3"].value == 20
    assert worksheet["E4"].value == 30


def test_task_list_and_download_use_latest_output(tmp_path) -> None:
    excel_path = tmp_path / "analysis.xlsx"
    _build_target_excel(excel_path)

    with excel_path.open("rb") as file_obj:
        create_response = client.post(
            "/workflow/excel-update/tasks",
            files={"file": ("analysis.xlsx", file_obj, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    task_id = create_response.json()["task_id"]
    client.post(
        f"/workflow/excel-update/tasks/{task_id}/operations",
        data={"user_prompt": "把这张清欠表 3 月实际产值更新一下"},
    )

    list_response = client.get("/workflow/excel-update/tasks")
    assert list_response.status_code == 200
    tasks = list_response.json()
    target_task = next(item for item in tasks if item["task_id"] == task_id)
    assert target_task["operation_count"] == 1
    assert target_task["latest_target_column"] == "3月实际产值"

    download_response = client.get(f"/workflow/excel-update/tasks/{task_id}/file")
    assert download_response.status_code == 200
    assert "step_01" in download_response.headers["content-disposition"]


def test_task_supports_excel_file_source_updates(tmp_path) -> None:
    target_excel_path = tmp_path / "target.xlsx"
    source_excel_path = tmp_path / "source.xlsx"
    _build_target_excel(target_excel_path)
    _build_source_excel(source_excel_path)

    with target_excel_path.open("rb") as file_obj:
        create_response = client.post(
            "/workflow/excel-update/tasks",
            files={"file": ("target.xlsx", file_obj, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert create_response.status_code == 200
    task_id = create_response.json()["task_id"]

    with source_excel_path.open("rb") as source_obj:
        operation_response = client.post(
            f"/workflow/excel-update/tasks/{task_id}/operations",
            files={"source_file": ("source.xlsx", source_obj, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={
                "source_type": "excel_file",
                "user_prompt": "使用源excel的项目编号、回款金额，来更新清欠表的三月回款",
            },
        )

    assert operation_response.status_code == 200
    payload = operation_response.json()
    assert payload["request"]["source_type"] == "excel_file"
    assert payload["request"]["source_match_column"] == "项目编号"
    assert payload["request"]["source_value_column"] == "回款金额"
    assert payload["request"]["target_column"] == "3月回款"
    assert payload["result"]["summary"]["updated_cells"] == 2

    workbook = load_workbook(payload["result"]["output_path"])
    worksheet = workbook["软件项目清欠表_收付款"]
    assert worksheet["E3"].value == 559336.13
    assert worksheet["E4"].value == 223344.55
