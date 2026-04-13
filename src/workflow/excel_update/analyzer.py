import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.config.settings import settings
from src.models.llm import chat_completion
from src.workflow.excel_update.schemas import (
    ExcelSheetAnalysis,
    ExcelUpdateAnalysisResult,
    ExcelUpdateQueryCondition,
)


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().split())


def _parse_json_object(text: str) -> dict[str, Any]:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(text[start : end + 1])
        raise


def _scan_sheet_headers(excel_path: str) -> list[ExcelSheetAnalysis]:
    workbook = load_workbook(Path(excel_path), read_only=True, data_only=False)
    sheet_options: list[ExcelSheetAnalysis] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        headers: list[str] = []
        seen: set[str] = set()
        max_scan_rows = min(worksheet.max_row, 10)
        for row in worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True):
            for cell_value in row:
                text = _normalize_text(cell_value)
                if not text or text in seen:
                    continue
                headers.append(text)
                seen.add(text)
        sheet_options.append(
            ExcelSheetAnalysis(
                sheet_name=sheet_name,
                header_candidates=headers,
            )
        )
    workbook.close()
    return sheet_options


def _infer_query_conditions(user_prompt: str) -> list[ExcelUpdateQueryCondition]:
    prompt = user_prompt.strip()
    if not prompt:
        return []

    now = datetime.now()
    year_match = re.search(r"(20\d{2})\s*年", prompt)
    month_match = re.search(r"(\d{1,2})\s*月", prompt)
    year = int(year_match.group(1)) if year_match else now.year

    conditions: list[ExcelUpdateQueryCondition] = []
    if month_match:
        month = int(month_match.group(1))
        conditions.append(
            ExcelUpdateQueryCondition(
                key="month",
                value=f"{year:04d}-{month:02d}",
            )
        )

    metric_map = {
        "实际产值": "actual_output",
        "产值": "actual_output",
        "回款": "receipts",
        "收付现": "cash_flow",
        "债权": "debt",
    }
    metric = next((value for keyword, value in metric_map.items() if keyword in prompt), None)
    if metric:
        conditions.append(ExcelUpdateQueryCondition(key="metric", value=metric))

    return conditions


def _preferred_target_names(user_prompt: str) -> list[str]:
    month_match = re.search(r"(\d{1,2})\s*月", user_prompt)
    metric_text = next(
        (
            keyword
            for keyword in ("实际产值", "回款", "收付现", "债权", "产值")
            if keyword in user_prompt
        ),
        "",
    )
    options: list[str] = []
    if month_match and metric_text:
        month = str(int(month_match.group(1)))
        options.extend([f"{month}月{metric_text}", f"{month}月份{metric_text}", f"三月{metric_text}" if month == "3" else ""])
    elif metric_text:
        options.append(metric_text)
    return [_normalize_text(option) for option in options if option]


def _score_sheet(sheet: ExcelSheetAnalysis, user_prompt: str) -> int:
    score = 0
    combined_text = f"{sheet.sheet_name} {' '.join(sheet.header_candidates)}"
    for keyword in ("清欠", "项目", "收付款", "回款", "产值", "债权"):
        if keyword in combined_text and keyword in user_prompt:
            score += 3
        elif keyword in combined_text:
            score += 1
    if "项目编号" in sheet.header_candidates:
        score += 3
    return score


def _find_target_column(sheet: ExcelSheetAnalysis, user_prompt: str) -> str | None:
    headers_by_normalized = {_normalize_text(header): header for header in sheet.header_candidates}
    for candidate in _preferred_target_names(user_prompt):
        if candidate in headers_by_normalized:
            return headers_by_normalized[candidate]

    month_match = re.search(r"(\d{1,2})\s*月", user_prompt)
    keywords = [keyword for keyword in ("实际产值", "回款", "收付现", "债权", "产值") if keyword in user_prompt]
    month_text = f"{int(month_match.group(1))}月" if month_match else ""

    for header in sheet.header_candidates:
        normalized = _normalize_text(header)
        if month_text and month_text not in normalized:
            continue
        if keywords and not any(keyword in normalized for keyword in keywords):
            continue
        return header
    return None


def _find_source_columns(source_sheet: ExcelSheetAnalysis, user_prompt: str) -> tuple[str | None, str | None]:
    headers = source_sheet.header_candidates
    normalized_headers = {_normalize_text(header): header for header in headers}

    source_match_column = next(
        (
            normalized_headers[key]
            for key in ("项目编号", "项目编码", "编号")
            if key in normalized_headers
        ),
        None,
    )

    prompt_columns = re.findall(r"源\s*excel.*?的(.+?)来更新", user_prompt)
    explicit_column_text = prompt_columns[0] if prompt_columns else ""
    explicit_parts = [
        part.strip(" ，、,")
        for part in re.split(r"[、,，和]\s*", explicit_column_text)
        if part.strip(" ，、,")
    ]

    if explicit_parts:
        for part in explicit_parts:
            normalized_part = _normalize_text(part)
            if source_match_column is None and normalized_part in normalized_headers:
                source_match_column = normalized_headers[normalized_part]
            elif normalized_part in normalized_headers:
                return source_match_column, normalized_headers[normalized_part]

    metric_candidates = []
    if "回款" in user_prompt:
        metric_candidates.extend(["回款金额", "回款", "收款金额"])
    if "实际产值" in user_prompt or "产值" in user_prompt:
        metric_candidates.extend(["实际产值", "产值", "完成产值"])

    source_value_column = next(
        (
            normalized_headers[key]
            for key in metric_candidates
            if key in normalized_headers
        ),
        None,
    )

    return source_match_column, source_value_column


def _heuristic_analysis(
    excel_path: str,
    user_prompt: str,
    source_excel_path: str | None = None,
) -> ExcelUpdateAnalysisResult:
    sheet_options = _scan_sheet_headers(excel_path)
    selected_sheet = max(sheet_options, key=lambda item: _score_sheet(item, user_prompt), default=None)
    warnings: list[str] = []

    sheet_name = selected_sheet.sheet_name if selected_sheet else None
    match_column = "项目编号"
    target_column = _find_target_column(selected_sheet, user_prompt) if selected_sheet else None

    if selected_sheet and match_column not in selected_sheet.header_candidates:
        similar = next((header for header in selected_sheet.header_candidates if "编号" in header), None)
        if similar:
            match_column = similar
        else:
            warnings.append("未在目标表头中明确识别到“项目编号”列，请确认匹配列。")

    if not target_column:
        warnings.append("未能可靠识别目标列，请确认目标列名。")

    source_sheet_options: list[ExcelSheetAnalysis] = []
    source_sheet_name: str | None = None
    source_match_column: str | None = None
    source_value_column: str | None = None
    if source_excel_path:
        source_sheet_options = _scan_sheet_headers(source_excel_path)
        selected_source_sheet = max(source_sheet_options, key=lambda item: _score_sheet(item, user_prompt), default=None)
        source_sheet_name = selected_source_sheet.sheet_name if selected_source_sheet else None
        if selected_source_sheet:
            source_match_column, source_value_column = _find_source_columns(selected_source_sheet, user_prompt)
        if not source_match_column:
            warnings.append("未能可靠识别源 Excel 的匹配列，请确认源表中的项目编号列。")
        if not source_value_column:
            warnings.append("未能可靠识别源 Excel 的取值列，请确认源表中的值列。")

    return ExcelUpdateAnalysisResult(
        user_prompt=user_prompt,
        sheet_name=sheet_name,
        match_column=match_column,
        match_field="project_no",
        source_sheet_name=source_sheet_name,
        source_match_column=source_match_column,
        source_value_column=source_value_column,
        target_column=target_column,
        query_conditions=_infer_query_conditions(user_prompt),
        sheet_options=sheet_options,
        source_sheet_options=source_sheet_options,
        warnings=warnings,
    )


def _llm_is_available() -> bool:
    return any(
        (
            os.getenv("OPENAI_API_KEY"),
            os.getenv("DASHSCOPE_API_KEY"),
            settings.dashscope_api_key,
        )
    )


def _refine_with_llm(base_result: ExcelUpdateAnalysisResult) -> ExcelUpdateAnalysisResult:
    response = chat_completion(
        temperature=0,
        max_tokens=700,
        messages=[
            {
                "role": "system",
                "content": (
                    "你是 Excel 更新参数识别器。"
                    "你只能返回 JSON 对象，不要输出解释。"
                    "字段必须包含 sheet_name, match_column, match_field, target_column, "
                    "source_sheet_name, source_match_column, source_value_column, query_conditions, warnings。"
                    "match_field 默认 project_no。"
                    "所有列名都只能从给定 sheet/header 候选中选择。"
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "user_prompt": base_result.user_prompt,
                        "target_sheet_options": [item.model_dump(mode="json") for item in base_result.sheet_options],
                        "source_sheet_options": [item.model_dump(mode="json") for item in base_result.source_sheet_options],
                        "heuristic_result": base_result.model_dump(mode="json"),
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
            },
        ],
    )
    payload = _parse_json_object(response)

    valid_target_sheet_names = {item.sheet_name for item in base_result.sheet_options}
    sheet_name = payload.get("sheet_name")
    if sheet_name not in valid_target_sheet_names:
        sheet_name = base_result.sheet_name

    selected_sheet = next((item for item in base_result.sheet_options if item.sheet_name == sheet_name), None)
    valid_headers = set(selected_sheet.header_candidates if selected_sheet else [])

    match_column = payload.get("match_column")
    if match_column not in valid_headers:
        match_column = base_result.match_column

    target_column = payload.get("target_column")
    if target_column not in valid_headers:
        target_column = base_result.target_column

    valid_source_sheet_names = {item.sheet_name for item in base_result.source_sheet_options}
    source_sheet_name = payload.get("source_sheet_name")
    if source_sheet_name not in valid_source_sheet_names:
        source_sheet_name = base_result.source_sheet_name

    selected_source_sheet = next(
        (item for item in base_result.source_sheet_options if item.sheet_name == source_sheet_name),
        None,
    )
    valid_source_headers = set(selected_source_sheet.header_candidates if selected_source_sheet else [])

    source_match_column = payload.get("source_match_column")
    if source_match_column not in valid_source_headers:
        source_match_column = base_result.source_match_column

    source_value_column = payload.get("source_value_column")
    if source_value_column not in valid_source_headers:
        source_value_column = base_result.source_value_column

    raw_conditions = payload.get("query_conditions") or []
    try:
        query_conditions = [
            ExcelUpdateQueryCondition.model_validate(item)
            for item in raw_conditions
            if isinstance(item, dict)
        ]
    except Exception:
        query_conditions = base_result.query_conditions

    warnings = [str(item) for item in payload.get("warnings", []) if str(item).strip()]
    return ExcelUpdateAnalysisResult(
        user_prompt=base_result.user_prompt,
        sheet_name=sheet_name,
        match_column=match_column,
        match_field=str(payload.get("match_field") or base_result.match_field),
        source_sheet_name=source_sheet_name,
        source_match_column=source_match_column,
        source_value_column=source_value_column,
        target_column=target_column,
        query_conditions=query_conditions or base_result.query_conditions,
        sheet_options=base_result.sheet_options,
        source_sheet_options=base_result.source_sheet_options,
        warnings=warnings or base_result.warnings,
    )


def analyze_excel_update(
    excel_path: str,
    user_prompt: str,
    source_excel_path: str | None = None,
) -> ExcelUpdateAnalysisResult:
    if not user_prompt.strip():
        raise ValueError("user_prompt must not be empty")

    base_result = _heuristic_analysis(
        excel_path=excel_path,
        user_prompt=user_prompt,
        source_excel_path=source_excel_path,
    )
    if not _llm_is_available():
        return base_result

    try:
        return _refine_with_llm(base_result)
    except Exception:
        return base_result
