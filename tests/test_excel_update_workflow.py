from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from src.workflow.excel_update.schemas import ExcelUpdateRequest
from src.workflow.excel_update.service import ExcelUpdateService


def _build_sample_excel(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "qingqian"
    worksheet.append(["项目编号", "项目名称", "3月实际产值"])
    worksheet.append(["HKZC-N-YW-2021-001", "项目A", None])
    worksheet.append(["HKZC-N-YW-2021-002", "项目B", 5])
    worksheet.append(["HKZC-N-YW-2021-003", "项目C", None])
    workbook.save(path)


def test_excel_update_service_writes_value_to_target_column(tmp_path) -> None:
    excel_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "updated.xlsx"
    _build_sample_excel(excel_path)

    request = ExcelUpdateRequest(
        excel_path=str(excel_path),
        sheet_name="qingqian",
        match_column="项目编号",
        match_field="project_no",
        target_column="3月实际产值",
        output_path=str(output_path),
    )

    result = ExcelUpdateService().run(request)

    assert result.summary.total_records == 2
    assert result.summary.matched_records == 2
    assert result.summary.updated_cells == 2
    assert result.summary.error_count == 0

    workbook = load_workbook(output_path)
    worksheet = workbook["qingqian"]
    assert worksheet["C2"].value == 20
    assert worksheet["C3"].value == 30


def test_excel_update_service_collects_unmatched_keys(tmp_path) -> None:
    excel_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "updated.xlsx"
    _build_sample_excel(excel_path)

    def custom_fetcher(_: ExcelUpdateRequest) -> list[dict[str, object]]:
        return [
            {"project_no": "HKZC-N-YW-2021-001", "value": 20},
            {"project_no": "HKZC-N-YW-2021-999", "value": 30},
        ]

    request = ExcelUpdateRequest(
        excel_path=str(excel_path),
        sheet_name="qingqian",
        match_column="项目编号",
        match_field="project_no",
        target_column="3月实际产值",
        output_path=str(output_path),
    )

    result = ExcelUpdateService(fetcher=custom_fetcher).run(request)

    assert result.unmatched_keys == ["HKZC-N-YW-2021-999"]
    assert result.summary.total_records == 2
    assert result.summary.matched_records == 1
    assert result.summary.unmatched_records == 1


def test_excel_update_service_supports_multiline_target_header(tmp_path) -> None:
    excel_path = tmp_path / "source_multiline.xlsx"
    output_path = tmp_path / "updated_multiline.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "qingqian"
    worksheet.append(["说明", "", "", ""])
    worksheet.append(["序号", "项目编号", "项目名称", "2026年回款、债权"])
    worksheet.append(["", "", "", "3月\n实际产值"])
    worksheet.append(["", "", "", ""])
    worksheet.append(["1", "HKZC-N-YW-2021-001", "项目A", None])
    worksheet.append(["2", "HKZC-N-YW-2021-002", "项目B", None])
    workbook.save(excel_path)

    request = ExcelUpdateRequest(
        excel_path=str(excel_path),
        sheet_name="qingqian",
        match_column="项目编号",
        match_field="project_no",
        target_column="3月实际产值",
        output_path=str(output_path),
    )

    result = ExcelUpdateService().run(request)

    assert result.summary.matched_records == 2
    assert result.summary.updated_cells == 2

    workbook = load_workbook(output_path)
    worksheet = workbook["qingqian"]
    assert worksheet["D5"].value == 20
    assert worksheet["D6"].value == 30


def test_excel_update_service_preserves_unmodified_zip_parts(tmp_path) -> None:
    excel_path = tmp_path / "source_with_parts.xlsx"
    output_path = tmp_path / "updated_with_parts.xlsx"
    _build_sample_excel(excel_path)

    augmented_path = tmp_path / "augmented.xlsx"
    extra_parts = {
        "xl/media/image99.png": b"fake-image-payload",
        "xl/cellimages.xml": b"<cellImages/>",
        "xl/_rels/cellimages.xml.rels": b"<Relationships/>",
        "xl/calcChain.xml": b"<calcChain/>",
    }
    with ZipFile(excel_path) as source_archive, ZipFile(augmented_path, "w") as target_archive:
        for info in source_archive.infolist():
            target_archive.writestr(info.filename, source_archive.read(info.filename), compress_type=info.compress_type)
        for name, payload in extra_parts.items():
            target_archive.writestr(name, payload, compress_type=ZIP_DEFLATED)

    request = ExcelUpdateRequest(
        excel_path=str(augmented_path),
        sheet_name="qingqian",
        match_column="项目编号",
        match_field="project_no",
        target_column="3月实际产值",
        output_path=str(output_path),
    )

    ExcelUpdateService().run(request)

    with ZipFile(output_path) as archive:
        for name, payload in extra_parts.items():
            assert archive.read(name) == payload


def test_excel_update_service_only_rewrites_target_sheet_xml(tmp_path) -> None:
    excel_path = tmp_path / "source_two_sheets.xlsx"
    output_path = tmp_path / "updated_two_sheets.xlsx"

    workbook = Workbook()
    target_sheet = workbook.active
    target_sheet.title = "qingqian"
    target_sheet.append(["项目编号", "项目名称", "3月实际产值"])
    target_sheet.append(["HKZC-N-YW-2021-001", "项目A", None])

    untouched_sheet = workbook.create_sheet("summary")
    untouched_sheet["A1"] = "keep-me"
    workbook.save(excel_path)

    with ZipFile(excel_path) as archive:
        before_summary_xml = archive.read("xl/worksheets/sheet2.xml")

    request = ExcelUpdateRequest(
        excel_path=str(excel_path),
        sheet_name="qingqian",
        match_column="项目编号",
        match_field="project_no",
        target_column="3月实际产值",
        output_path=str(output_path),
    )

    ExcelUpdateService().run(request)

    with ZipFile(output_path) as archive:
        after_summary_xml = archive.read("xl/worksheets/sheet2.xml")

    assert before_summary_xml == after_summary_xml
