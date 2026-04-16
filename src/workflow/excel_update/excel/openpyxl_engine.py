from __future__ import annotations

from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from zipfile import ZIP_DEFLATED, ZipFile

from src.workflow.excel_update.schemas import OperationExecutionResult, OperationItemResultData, OperationSummaryData


class OpenpyxlExcelEngine:
    def list_sheet_names(self, payload: bytes) -> list[str]:
        workbook = _load_workbook(filename=BytesIO(payload), read_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def apply_updates(
        self,
        *,
        input_path: Path,
        output_path: Path,
        sheet_name: str,
        match_column: str,
        target_column: str,
        updates: list[dict],
    ) -> OperationExecutionResult:
        workbook = _load_workbook(filename=input_path)
        worksheet = workbook[sheet_name]
        rows_by_value = self._build_row_index(worksheet=worksheet, match_column=match_column)

        items: list[OperationItemResultData] = []
        success_count = 0
        not_found_count = 0
        duplicate_count = 0
        failed_count = 0

        for update in updates:
            match_value = str(update["match_value"]).strip()
            target_value = update.get("target_value")
            rows = rows_by_value.get(match_value, [])
            if not rows:
                not_found_count += 1
                items.append(
                    OperationItemResultData(
                        match_value=match_value,
                        target_value=target_value,
                        status="not_found",
                        message="no matched row found",
                    )
                )
                continue
            if len(rows) > 1:
                duplicate_count += 1
                items.append(
                    OperationItemResultData(
                        match_value=match_value,
                        target_value=target_value,
                        status="duplicate_match",
                        message="multiple matched rows found",
                    )
                )
                continue

            row_index = rows[0]
            cell_address = f"{target_column}{row_index}"
            cell = worksheet[cell_address]
            old_value = cell.value
            try:
                cell.value = target_value
            except Exception as exc:
                failed_count += 1
                items.append(
                    OperationItemResultData(
                        match_value=match_value,
                        target_value=target_value,
                        status="failed",
                        row_index=row_index,
                        cell_address=cell_address,
                        old_value=self._stringify_value(old_value),
                        message=str(exc),
                    )
                )
                continue

            success_count += 1
            items.append(
                OperationItemResultData(
                    match_value=match_value,
                    target_value=target_value,
                    status="success",
                    row_index=row_index,
                    cell_address=cell_address,
                    old_value=self._stringify_value(old_value),
                    new_value=self._stringify_value(cell.value),
                )
            )

        if success_count > 0:
            target_sheet_index = workbook.sheetnames.index(sheet_name) + 1
            total_sheet_count = len(workbook.sheetnames)
            workbook.save(output_path)
            workbook.close()
            self._preserve_unmodified_members(
                original_path=input_path,
                updated_path=output_path,
                final_path=output_path,
                target_sheet_index=target_sheet_index,
                total_sheet_count=total_sheet_count,
            )
        else:
            workbook.close()

        summary = OperationSummaryData(
            total_count=len(updates),
            success_count=success_count,
            not_found_count=not_found_count,
            duplicate_count=duplicate_count,
            failed_count=failed_count,
        )
        return OperationExecutionResult(
            summary=summary,
            items=items,
            has_successful_updates=success_count > 0,
        )

    def _build_row_index(self, *, worksheet, match_column: str) -> dict[str, list[int]]:
        rows_by_value: dict[str, list[int]] = {}
        for row_index in range(1, worksheet.max_row + 1):
            raw_value = worksheet[f"{match_column}{row_index}"].value
            if raw_value is None:
                continue
            normalized = str(raw_value).strip()
            if not normalized:
                continue
            rows_by_value.setdefault(normalized, []).append(row_index)
        return rows_by_value

    def _preserve_unmodified_members(
        self,
        *,
        original_path: Path,
        updated_path: Path,
        final_path: Path,
        target_sheet_index: int,
        total_sheet_count: int,
    ) -> None:
        preserved_names = {
            f"xl/worksheets/sheet{sheet_index}.xml"
            for sheet_index in range(1, total_sheet_count + 1)
            if sheet_index != target_sheet_index
        }
        preserved_names.update(
            {
                f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
                for sheet_index in range(1, total_sheet_count + 1)
                if sheet_index != target_sheet_index
            }
        )

        with (
            ZipFile(original_path) as original_archive,
            ZipFile(updated_path) as updated_archive,
            NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file,
        ):
            temp_path = Path(temp_file.name)

        with ZipFile(original_path) as original_archive, ZipFile(updated_path) as updated_archive, ZipFile(
            temp_path, "w"
        ) as final_archive:
            original_infos = {info.filename: info for info in original_archive.infolist()}
            updated_infos = {info.filename: info for info in updated_archive.infolist()}

            for name, info in updated_infos.items():
                source_archive = original_archive if name in preserved_names and name in original_infos else updated_archive
                source_info = original_infos.get(name, info)
                final_archive.writestr(
                    info.filename,
                    source_archive.read(name),
                    compress_type=source_info.compress_type or ZIP_DEFLATED,
                )

            for name, info in original_infos.items():
                if name in updated_infos:
                    continue
                final_archive.writestr(
                    name,
                    original_archive.read(name),
                    compress_type=info.compress_type or ZIP_DEFLATED,
                )

        temp_path.replace(final_path)

    @staticmethod
    def _stringify_value(value) -> str | None:
        if value is None:
            return None
        return str(value)


def _load_workbook(*args, **kwargs):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for excel update operations") from exc
    return load_workbook(*args, **kwargs)
