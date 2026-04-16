from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _build_workbook(path: Path) -> None:
    workbook = Workbook()
    cover_sheet = workbook.active
    cover_sheet.title = "封面"
    cover_sheet["A1"] = "说明"

    data_sheet = workbook.create_sheet("3月产值")
    data_sheet.append(["项目编号", "项目名称", "回款"])
    data_sheet.append(["P001", "项目A", 10])
    data_sheet.append(["P002", "项目B", 20])
    workbook.save(path)


def _relative_path(url: str) -> str:
    return url.replace("http://testserver", "")


def test_upload_task_and_get_detail(client, excel_env: Path) -> None:
    excel_path = excel_env / "upload.xlsx"
    _build_workbook(excel_path)

    with excel_path.open("rb") as file_obj:
        response = client.post(
            "/api/excel-tasks/upload",
            files={
                "file": (
                    "upload.xlsx",
                    file_obj,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"task_name": "3月产值更新", "created_by": "zhangsan"},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 0
    task_id = payload["data"]["task_id"]
    assert payload["data"]["task_name"] == "3月产值更新"
    assert payload["data"]["current_version_no"] == 1

    list_response = client.get("/api/excel-tasks")
    assert list_response.status_code == 200
    list_payload = list_response.json()
    assert list_payload["data"]["total"] == 1
    assert list_payload["data"]["items"][0]["task_id"] == task_id
    assert list_payload["data"]["items"][0]["operation_count"] == 0

    detail_response = client.get(f"/api/excel-tasks/{task_id}")
    assert detail_response.status_code == 200
    detail_payload = detail_response.json()
    assert detail_payload["data"]["sheet_names"] == ["封面", "3月产值"]
    assert detail_payload["data"]["created_by"] == "zhangsan"
    assert detail_payload["data"]["latest_version"]["version_no"] == 1


def test_original_download_returns_uploaded_file(client, excel_env: Path) -> None:
    excel_path = excel_env / "original.xlsx"
    _build_workbook(excel_path)

    with excel_path.open("rb") as file_obj:
        create_response = client.post(
            "/api/excel-tasks/upload",
            files={
                "file": (
                    "original.xlsx",
                    file_obj,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    task_id = create_response.json()["data"]["task_id"]
    download_response = client.get(f"/api/excel-tasks/{task_id}/download/original")
    assert download_response.status_code == 200
    download_url = download_response.json()["data"]["download_url"]

    file_response = client.get(_relative_path(download_url))
    assert file_response.status_code == 200
    assert "attachment;" in file_response.headers["content-disposition"]

    workbook = load_workbook(filename=BytesIO(file_response.content))
    assert workbook.sheetnames == ["封面", "3月产值"]
    worksheet = workbook["3月产值"]
    assert worksheet["C2"].value == 10
