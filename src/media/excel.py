from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from src.media.schemas import FrameInspectionResult, VideoInspectionResult

REPORT_SHEET_NAME = "巡检结果"
REPORT_HEADERS = [
    "视频文件",
    "抽帧间隔(秒)",
    "帧序号",
    "时间点",
    "截图",
    "安全帽颜色",
    "反光衣颜色",
    "爆闪灯颜色",
    "是否为管理人员",
    "判定依据",
    "原始输出",
]


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _format_timestamp(seconds: int) -> str:
    hours, remain = divmod(seconds, 3600)
    minutes, secs = divmod(remain, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def _apply_layout(worksheet) -> None:
    widths = {
        "A": 28,
        "B": 12,
        "C": 10,
        "D": 12,
        "E": 34,
        "F": 16,
        "G": 18,
        "H": 16,
        "I": 14,
        "J": 56,
        "K": 56,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


class ExcelReportWriter:
    def __init__(self, *, video_path: str, interval_seconds: int, output_path: str) -> None:
        self.video_path = video_path
        self.interval_seconds = interval_seconds
        self.output_path = Path(output_path).expanduser()
        _ensure_parent(self.output_path)

        if self.output_path.is_file():
            self.workbook = load_workbook(self.output_path)
            self.worksheet = self.workbook[REPORT_SHEET_NAME]
            self._row_index = self.worksheet.max_row + 1
        else:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = REPORT_SHEET_NAME
            self.worksheet.append(REPORT_HEADERS)
            _apply_layout(self.worksheet)
            self.worksheet.freeze_panes = "A2"
            self._row_index = 2
            self.save()

    def append_frame(self, frame: FrameInspectionResult) -> None:
        parsed = frame.parsed_result or {}
        self.worksheet.append(
            [
                self.video_path,
                self.interval_seconds,
                frame.frame_index,
                _format_timestamp(frame.timestamp_seconds),
                "",
                parsed.get("安全帽颜色", ""),
                parsed.get("反光衣颜色", ""),
                parsed.get("爆闪灯颜色", ""),
                parsed.get("是否为管理人员", ""),
                parsed.get("判定依据", ""),
                frame.raw_answer,
            ]
        )

        self.worksheet.row_dimensions[self._row_index].height = 180
        for column_index in range(1, len(REPORT_HEADERS) + 1):
            cell = self.worksheet.cell(row=self._row_index, column=column_index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if frame.frame_path and Path(frame.frame_path).is_file():
            image = XLImage(frame.frame_path)
            image.width = 240
            image.height = 135
            self.worksheet.add_image(image, f"E{self._row_index}")

        self._row_index += 1
        self.save()

    def save(self) -> str:
        self.workbook.save(self.output_path)
        return str(self.output_path)


def export_video_inspection_report(
    result: VideoInspectionResult,
    output_path: str,
) -> str:
    writer = ExcelReportWriter(
        video_path=result.video_path,
        interval_seconds=result.interval_seconds,
        output_path=output_path,
    )
    for frame in result.frames:
        writer.append_frame(frame)
    return str(writer.output_path)
